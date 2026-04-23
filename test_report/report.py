#!/usr/bin/env python3
"""
WorkLogger weekly report generator — budget-fill pipeline.

Pipeline:
  1. Load all JSONL events for the week
  2. Split into work slots (idle/sleep/lock/gap boundaries)
  3. Aggregate app + project time per slot
  4. Fetch ALL git commits for each calendar day
  5. Harvest deduplicated work packages from aggregated data
  6. Budget-fill: place immovable items, then pack work packages into free time

Usage:
  python3 test_report/report.py              # current ISO week
  python3 test_report/report.py 2026 16      # specific year + week number
  python3 test_report/report.py --csv        # write report.xlsx instead of printing
"""

import json, subprocess, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

# ── Load config ───────────────────────────────────────────────────────────────

def _find_config() -> Path:
    candidates = [
        # User config written by the app's Settings (highest priority)
        Path.home() / "Library/Application Support/WorkLogger/config.json",
        # Repo root (dev workflow)
        Path(__file__).parent.parent / "config.json",
        # App bundle fallback
        Path.home() / "Desktop/WorkLogger.app/Contents/Resources/config.json",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("config.json not found")

_cfg     = json.loads(_find_config().read_text())
_rep     = _cfg.get("report", {})

LOG_DIR          = Path(_cfg["logDirectory"])
REPOS            = [Path(p) for p in _rep.get("repositories", [])]
GAP_MINUTES      = _rep.get("gapMinutes",      10)
MIN_SLOT_MINUTES = _rep.get("minSlotMinutes",    5)
BLOCK_MINUTES    = _rep.get("blockMinutes",     30)
DAY_HOURS        = _rep.get("dayHours",          8)
LUNCH_MINUTES    = _rep.get("lunchMinutes",     60)
LUNCH_START      = _rep.get("lunchStart",   "12:00")
MIN_PACKAGE_MIN  = _rep.get("minPackageMinutes", 15)
SKIP_APPS        = {
    "Code", "Finder", "loginwindow", "universalAccessAuthWarn",
    "UserNotificationCenter", "WorkLogger", "Terminal",
    "Safari", "Microsoft Teams",
} | set(_rep.get("skipApps", []))
# Hardcoded noise that is always filtered regardless of user config
_BUILTIN_SAFARI_EXACT = {
    "favorites://", "Personal Desktop", "Start Page", "502 Bad Gateway", "Untitled",
    "Just a moment...", "Failed to open page", "reading-list://",
}
_BUILTIN_SAFARI_CONTAINS = {
    "#code=", "spa-signin-oidc", "access_token=", "?code=", "session_state=",
    "#loginHint=", "windows.cloud.microsoft", "Sign in to your account", "Windows App",
    "_MsalRedirect", "/_msal/", "vssps.visualstudio.com",
}
SKIP_SAFARI_EXACT    = _BUILTIN_SAFARI_EXACT | set(_rep.get("skipSafariExact", []))
SKIP_SAFARI_CONTAINS = _BUILTIN_SAFARI_CONTAINS | set(_rep.get("skipSafariContains", []))
SHOW_SAFARI_TIME = _cfg.get("showSafariTimeInReport", False)

# Planned meeting slots per weekday: list of (start_time, end_time, description)
_raw_prefilled = _rep.get("prefilledSlots", {})
PREFILLED_SLOTS: dict[str, list[tuple[str, str, str]]] = {}
for _dow, _ranges in _raw_prefilled.items():
    if _dow == "comment":
        continue
    parsed = []
    for _r in _ranges:
        if isinstance(_r, dict):
            _time = _r.get("time", "")
            _desc = _r.get("description", "")
        else:
            _time = _r
            _desc = ""
        _s, _e = _time.split("-")
        parsed.append((_s.strip(), _e.strip(), _desc))
    PREFILLED_SLOTS[_dow] = parsed

BREAK_START = {"system_sleep", "screen_lock", "idle_start"}
BREAK_END   = {"system_wake",  "screen_unlock", "idle_end"}
NOISE       = {"permissions"}

# ── Timestamp parsing ─────────────────────────────────────────────────────────

def parse_ts(s: str) -> datetime:
    """Parse timestamps from any format WorkLogger emits, return naive local time."""
    s = s.strip()
    if s.endswith("Z"):
        dt = datetime.fromisoformat(s[:-1]).replace(tzinfo=timezone.utc)
        return dt.astimezone().replace(tzinfo=None)
    # git: "2026-04-15 12:00:00 +0200"  (space, no colon in offset)
    if " " in s and (s[-5] in "+-") and (":" not in s[-5:]):
        s = s[:-2] + ":" + s[-2:]
    s = s.replace(" ", "T", 1)
    dt = datetime.fromisoformat(s)
    if dt.tzinfo:
        dt = dt.astimezone().replace(tzinfo=None)
    return dt

# ── Load events ───────────────────────────────────────────────────────────────

def load_week(year: int, week: int) -> list[dict]:
    events = []
    for path in sorted(LOG_DIR.glob("*.jsonl")):
        try:
            date = datetime.strptime(path.stem, "%Y-%m-%d").date()
        except ValueError:
            continue
        if date.isocalendar()[:2] == (year, week):
            with open(path) as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        ev = json.loads(line)
                        ev["_date"] = path.stem
                        ev["_ts"]   = parse_ts(ev["timestamp"])
                        events.append(ev)
                    except Exception:
                        pass
    events.sort(key=lambda e: e["_ts"])
    return events

# ── Manual entries ───────────────────────────────────────────────────────────

def extract_manual_entries(events: list[dict]) -> list[dict]:
    """Pull all manual_entry events out of the event list.
    Each entry has description, time (HH:MM), duration_minutes, _date, _ts.
    """
    entries = []
    for ev in events:
        if ev.get("event") != "manual_entry":
            continue
        desc = ev.get("description", "").strip()
        if not desc:
            continue
        time_str  = ev.get("time", "")       # "HH:MM" as user entered
        dur_min   = int(ev.get("duration_minutes", 60))
        date_str  = ev["_date"]               # YYYY-MM-DD
        date      = datetime.strptime(date_str, "%Y-%m-%d")

        # Resolve start time: use user-supplied time field if valid, else use _ts
        try:
            h, m = map(int, time_str.split(":"))
            start = date.replace(hour=h, minute=m, second=0, microsecond=0)
        except Exception:
            start = ev["_ts"].replace(second=0, microsecond=0)

        end = start + timedelta(minutes=dur_min)
        entries.append({
            "date":     date_str,
            "start":    start,
            "end":      end,
            "dur_min":  dur_min,
            "description": desc,
            "_ts":      ev["_ts"],
        })
    return entries

# ── Slot detection ────────────────────────────────────────────────────────────

def split_slots(events: list[dict]) -> list[list[dict]]:
    slots, current, in_break = [], [], False
    for ev in events:
        evt = ev.get("event", "")
        if evt in NOISE:
            current.append(ev); continue
        if evt == "started" and current:
            slots.append(current); current = [ev]; in_break = False; continue
        if evt in BREAK_START:
            if current: current.append(ev); slots.append(current)
            current = []; in_break = True; continue
        if evt in BREAK_END:
            in_break = False; continue
        if current and not in_break:
            if (ev["_ts"] - current[-1]["_ts"]).total_seconds() / 60 > GAP_MINUTES:
                slots.append(current); current = [ev]; continue
        current.append(ev)
    if current:
        slots.append(current)
    return slots

# ── Aggregation per slot ──────────────────────────────────────────────────────

def aggregate(slot: list[dict]) -> dict | None:
    real = [e for e in slot if e.get("event") not in NOISE]
    if not real:
        return None
    t0, t1 = real[0]["_ts"], real[-1]["_ts"]
    dur = (t1 - t0).total_seconds() / 60
    if dur < MIN_SLOT_MINUTES:
        return None

    app_sec:     dict[str, float] = defaultdict(float)
    proj_sec:    dict[str, float] = defaultdict(float)
    safari_sec:  dict[str, float] = defaultdict(float)
    teams_seen:  list[str] = []
    safari_seen: list[str] = []
    manual_entries: list[dict] = []

    cur_app = cur_app_ts = None
    cur_proj = cur_proj_ts = None
    cur_safari = cur_safari_ts = None

    for i, ev in enumerate(real):
        evt = ev.get("event", "")
        ts  = ev["_ts"]

        if evt == "app_switch":
            if cur_app and cur_app_ts:
                app_sec[cur_app] += (ts - cur_app_ts).total_seconds()
            # close vscode project if leaving VS Code
            if cur_app not in ("Code", None) and cur_proj and cur_proj_ts:
                proj_sec[cur_proj] += (ts - cur_proj_ts).total_seconds()
                cur_proj = cur_proj_ts = None
            cur_app, cur_app_ts = ev.get("app", "?"), ts

            app = cur_app
            detail = ev.get("detail", "")

            # Use app_switch detail as project hint when VS Code is activated
            # and no explicit vscode_project_change has been seen yet
            if app == "Code" and detail and detail not in ("unknown window", ""):
                if cur_proj != detail:
                    # close previous project interval
                    if cur_proj and cur_proj_ts:
                        proj_sec[cur_proj] += (ts - cur_proj_ts).total_seconds()
                    cur_proj, cur_proj_ts = detail, ts

            if app == "Microsoft Teams" and detail:
                if detail not in teams_seen:
                    teams_seen.append(detail)

        elif evt == "vscode_project_change":
            if cur_proj and cur_proj_ts:
                proj_sec[cur_proj] += (ts - cur_proj_ts).total_seconds()
            cur_proj, cur_proj_ts = ev.get("detail", "?"), ts

        elif evt == "manual_entry":
            manual_entries.append({
                "description":      ev.get("description", ""),
                "time":             ev.get("time", ""),
                "duration_minutes": ev.get("duration_minutes", 60),
            })

        elif evt == "safari_tab_change":
            label = ev.get("detail") or ev.get("url", "")
            if label and label not in safari_seen:
                safari_seen.append(label)
            # Accumulate time per Safari tab
            if cur_safari and cur_safari_ts:
                safari_sec[cur_safari] += (ts - cur_safari_ts).total_seconds()
            cur_safari, cur_safari_ts = label, ts

    # close final intervals
    if cur_app and cur_app_ts:
        app_sec[cur_app] += (t1 - cur_app_ts).total_seconds()
    if cur_proj and cur_proj_ts:
        proj_sec[cur_proj] += (t1 - cur_proj_ts).total_seconds()
    if cur_safari and cur_safari_ts:
        safari_sec[cur_safari] += (t1 - cur_safari_ts).total_seconds()

    return {
        "date":          real[0]["_date"],
        "t0":            t0,
        "t1":            t1,
        "dur_min":       dur,
        "app_sec":       dict(sorted(app_sec.items(),  key=lambda x: -x[1])),
        "proj_sec":      dict(sorted(proj_sec.items(), key=lambda x: -x[1])),
        "safari_sec":    dict(sorted(safari_sec.items(), key=lambda x: -x[1])),
        "teams":         teams_seen,
        "safari":        safari_seen,
        "manual_entries": manual_entries,
    }

# ── Git commits ───────────────────────────────────────────────────────────────

def commits_for_day(repo: Path, date_str: str) -> list[dict]:
    day     = datetime.strptime(date_str, "%Y-%m-%d")
    after   = day.strftime("%Y-%m-%d 00:00:00")
    before  = day.strftime("%Y-%m-%d 23:59:59")
    try:
        r = subprocess.run(
            ["git", "-C", str(repo), "log",
             f"--after={after}", f"--before={before}",
             "--format=%H|%ai|%s", "--all"],
            capture_output=True, text=True, timeout=10
        )
    except Exception:
        return []
    out = []
    for line in r.stdout.strip().splitlines():
        parts = line.split("|", 2)
        if len(parts) != 3:
            continue
        sha, ts_raw, msg = parts
        try:
            ts = parse_ts(ts_raw)
        except Exception:
            continue
        out.append({"sha": sha[:8], "ts": ts, "msg": msg, "repo": repo.name})
    return out

def all_commits_for_dates(dates: list[str]) -> list[dict]:
    commits = []
    for date in dates:
        for repo in REPOS:
            commits.extend(commits_for_day(repo, date))
    commits.sort(key=lambda c: c["ts"])
    return commits

# ── Attach commits to aggs ────────────────────────────────────────────────────

def attach_commits(aggs: list[dict], all_commits: list[dict]) -> None:
    """Attach commits to aggregated slots for package harvesting."""
    for agg in aggs:
        agg["commits"] = []
        agg.setdefault("manual_entries", [])

    for commit in all_commits:
        ct = commit["ts"]
        date_str = ct.strftime("%Y-%m-%d")
        matched = next((a for a in aggs if a["date"] == date_str and a["t0"] <= ct <= a["t1"]), None)
        if matched is None:
            same_day = [a for a in aggs if a["date"] == date_str]
            if same_day:
                matched = min(same_day, key=lambda a: abs(((a["t0"] + (a["t1"] - a["t0"]) / 2) - ct).total_seconds()))
        if matched:
            matched["commits"].append(commit)

# ── Round to grid ─────────────────────────────────────────────────────────────

def round_down(dt: datetime, minutes: int) -> datetime:
    return dt - timedelta(minutes=dt.minute % minutes, seconds=dt.second, microseconds=dt.microsecond)

def round_up(dt: datetime, minutes: int) -> datetime:
    excess = timedelta(minutes=dt.minute % minutes, seconds=dt.second, microseconds=dt.microsecond)
    return (dt - excess + timedelta(minutes=minutes)) if excess else dt

def _fmt(sec: float) -> str:
    m = int(sec / 60)
    return f"{m//60}h{m%60:02d}m" if m >= 60 else f"{m}min"

# ── Work package harvesting ───────────────────────────────────────────────────

WEEKDAYS_DE = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]
WEEKDAYS_EN = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

def _prefilled_blocks(date: datetime) -> list[tuple[datetime, datetime, str]]:
    """Return planned meeting windows for a given date as (start, end, description)."""
    dow_en = WEEKDAYS_EN[date.weekday()]
    blocks = []
    for start_str, end_str, desc in PREFILLED_SLOTS.get(dow_en, []):
        sh, sm = map(int, start_str.split(":"))
        eh, em = map(int, end_str.split(":"))
        blocks.append((
            date.replace(hour=sh, minute=sm, second=0, microsecond=0),
            date.replace(hour=eh, minute=em, second=0, microsecond=0),
            desc,
        ))
    return blocks


def _extract_meeting_name(title: str) -> str | None:
    """Extract meaningful name from a Teams call window title."""
    if "Kompakte Besprechungsansicht" not in title:
        return None
    segments = [s.strip() for s in title.split("|")]
    meaningful = [s for s in segments if s and s not in
                  ("Microsoft Teams", "Calendar", "Kompakte Besprechungsansicht",
                   "Chat", "Calendar | Calendar")]
    return meaningful[0] if meaningful else None


def harvest_packages(aggs: list[dict], manual_entries: list[dict]) -> dict[str, list[dict]]:
    """Extract deduplicated work packages per date from all sources.

    Returns: {date_str: [package, ...]} where each package is:
      {"type": str, "description": str, "weight": float (seconds),
       "source_t0": datetime, "source_t1": datetime}
    """
    packages: dict[str, list[dict]] = defaultdict(list)

    for agg in aggs:
        date_str = agg["date"]

        # Track which project time is "claimed" by commits
        claimed_projects: set[str] = set()

        # ── Teams calls (detected via window titles) ─────────────────────
        call_names_seen: set[str] = set()
        for title in agg.get("teams", []):
            name = _extract_meeting_name(title)
            if name and name not in call_names_seen:
                call_names_seen.add(name)
                # Estimate call duration from Teams app time, divided by call count
                teams_calls = [t for t in agg["teams"] if _extract_meeting_name(t)]
                per_call = agg["app_sec"].get("Microsoft Teams", 0) / max(1, len(teams_calls))
                weight = max(per_call, 15 * 60)  # at least 15 min
                packages[date_str].append({
                    "type": "call",
                    "description": f"Call: {name}",
                    "weight": weight,
                    "source_t0": agg["t0"],
                    "source_t1": agg["t1"],
                })

        # ── Git commits (grouped by repo) ────────────────────────────────
        commits_by_repo: dict[str, list[str]] = {}
        for c in agg.get("commits", []):
            commits_by_repo.setdefault(c["repo"], []).append(c["msg"])
        for repo, msgs in commits_by_repo.items():
            proj_time = agg["proj_sec"].get(repo, 0)
            weight = max(proj_time, len(msgs) * 5 * 60)  # at least 5min per commit
            claimed_projects.add(repo)
            commit_desc = ", ".join(msgs)
            packages[date_str].append({
                "type": "project",
                "description": f"[{repo}] {commit_desc}",
                "weight": weight,
                "source_t0": agg["t0"],
                "source_t1": agg["t1"],
            })

        # ── VS Code projects (not already claimed by commits) ────────────
        for proj, sec in agg["proj_sec"].items():
            if proj in claimed_projects:
                continue
            if sec < 60 or proj in ("Save As", "unknown"):
                continue
            packages[date_str].append({
                "type": "project",
                "description": proj,
                "weight": sec,
                "source_t0": agg["t0"],
                "source_t1": agg["t1"],
            })

        # ── Apps (non-noise, non-code, non-teams) ────────────────────────
        for app, sec in agg["app_sec"].items():
            if app in SKIP_APPS or sec < 120:
                continue
            packages[date_str].append({
                "type": "app",
                "description": app,
                "weight": sec,
                "source_t0": agg["t0"],
                "source_t1": agg["t1"],
            })

        # ── Safari / Web research ────────────────────────────────────────
        web_tabs = []
        for tab in agg.get("safari", []):
            if not tab or tab in SKIP_SAFARI_EXACT:
                continue
            if any(s in tab for s in SKIP_SAFARI_CONTAINS):
                continue
            web_tabs.append(tab)
        if web_tabs:
            total_safari = sum(
                sec for tab, sec in agg.get("safari_sec", {}).items()
                if tab in web_tabs
            )
            if total_safari < 60:
                total_safari = 60
            desc_tabs = web_tabs[:5]
            packages[date_str].append({
                "type": "web_research",
                "description": "Web: " + ", ".join(desc_tabs),
                "weight": total_safari,
                "source_t0": agg["t0"],
                "source_t1": agg["t1"],
            })

    # ── Manual entries ───────────────────────────────────────────────────
    for entry in manual_entries:
        packages[entry["date"]].append({
            "type": "manual",
            "description": entry["description"],
            "weight": entry["dur_min"] * 60,
            "source_t0": entry["start"],
            "source_t1": entry["end"],
        })

    # ── Deduplicate within each day ──────────────────────────────────────
    for date_str in packages:
        packages[date_str] = _dedup_packages(packages[date_str])

    return dict(packages)


def _dedup_packages(pkgs: list[dict]) -> list[dict]:
    """Merge packages with identical descriptions, summing weights."""
    seen: dict[str, dict] = {}
    for pkg in pkgs:
        key = pkg["description"]
        if key in seen:
            existing = seen[key]
            existing["weight"] += pkg["weight"]
            existing["source_t0"] = min(existing["source_t0"], pkg["source_t0"])
            existing["source_t1"] = max(existing["source_t1"], pkg["source_t1"])
        else:
            seen[key] = dict(pkg)
    return sorted(seen.values(), key=lambda p: p["source_t0"])


# ── Budget-fill row placement ────────────────────────────────────────────────

def _parse_hm(s: str) -> tuple[int, int]:
    h, m = map(int, s.split(":"))
    return h, m


def make_rows(aggs: list[dict], manual_entries: list[dict] = [],
              packages_by_date: dict[str, list[dict]] | None = None) -> list[dict]:
    """
    Budget-fill algorithm:
      1. Place immovable items (planned meetings, lunch)
      2. Place detected Teams calls at their real time
      3. Proportionally fill remaining free time with work packages
    Each package appears exactly once — no duplication.
    """
    rows: list[dict] = []

    # Gather all active dates
    active_dates: set[str] = set()
    for agg in aggs:
        active_dates.add(agg["date"])
    for entry in manual_entries:
        active_dates.add(entry["date"])
    if packages_by_date:
        active_dates |= set(packages_by_date.keys())

    for date_str in sorted(active_dates):
        date = datetime.strptime(date_str, "%Y-%m-%d")
        dow_de = WEEKDAYS_DE[date.weekday()]

        # ── Determine work day boundaries from actual events ─────────────
        day_aggs = [a for a in aggs if a["date"] == date_str]
        day_manuals = [e for e in manual_entries if e["date"] == date_str]

        timestamps = ([a["t0"] for a in day_aggs] + [a["t1"] for a in day_aggs] +
                      [e["start"] for e in day_manuals])
        if not timestamps:
            continue

        day_start = round_down(min(timestamps), MIN_PACKAGE_MIN)
        available_min = DAY_HOURS * 60
        lh, lm = _parse_hm(LUNCH_START)
        lunch_s = date.replace(hour=lh, minute=lm, second=0, microsecond=0)
        lunch_e = lunch_s + timedelta(minutes=LUNCH_MINUTES)
        day_end = day_start + timedelta(minutes=available_min + LUNCH_MINUTES)

        # ── Reserved windows (immovable) ─────────────────────────────────
        reserved: list[tuple[datetime, datetime, str]] = []

        # Planned meetings
        for ps, pe, desc in _prefilled_blocks(date):
            reserved.append((ps, pe, desc if desc else "Meeting"))

        # Lunch break
        reserved.append((lunch_s, lunch_e, "Mittagspause"))

        # Sort reserved by start time
        reserved.sort(key=lambda x: x[0])

        # Emit reserved rows (skip lunch — it's just a gap)
        for rs, re, rdesc in reserved:
            if rdesc == "Mittagspause":
                continue
            rows.append({
                "Wochentag":    dow_de,
                "Datum":        date.strftime("%d.%m.%Y"),
                "Von":          rs.strftime("%H:%M"),
                "Bis":          re.strftime("%H:%M"),
                "Beschreibung": rdesc,
                "_start":       rs,
            })

        reserved_windows = [(s, e) for s, e, _ in reserved]

        # ── Compute free windows ─────────────────────────────────────────
        free_windows = _compute_free_windows(day_start, day_end, reserved_windows)

        total_free_min = sum(
            (fe - fs).total_seconds() / 60 for fs, fe in free_windows
        )

        # ── Get work packages for this day ───────────────────────────────
        day_pkgs = list(packages_by_date.get(date_str, [])) if packages_by_date else []

        # Separate calls (placed at real time) from fillable packages
        call_pkgs = [p for p in day_pkgs if p["type"] == "call"]
        fill_pkgs = [p for p in day_pkgs if p["type"] != "call"]

        # ── Place calls at their detected time, clipped to free windows ──
        for pkg in call_pkgs:
            call_dur = min(pkg["weight"] / 60, total_free_min)
            call_dur = max(call_dur, MIN_PACKAGE_MIN)
            # Try to place at source_t0, snapped to grid
            call_start = round_down(pkg["source_t0"], MIN_PACKAGE_MIN)
            call_end = call_start + timedelta(minutes=call_dur)

            # Clip to free windows
            placed = False
            for fi, (fs, fe) in enumerate(free_windows):
                if call_start >= fs and call_start < fe:
                    call_end = min(call_end, fe)
                    if (call_end - call_start).total_seconds() / 60 >= MIN_PACKAGE_MIN:
                        rows.append({
                            "Wochentag":    dow_de,
                            "Datum":        date.strftime("%d.%m.%Y"),
                            "Von":          call_start.strftime("%H:%M"),
                            "Bis":          call_end.strftime("%H:%M"),
                            "Beschreibung": pkg["description"],
                            "_start":       call_start,
                        })
                        # Split the free window around this call
                        free_windows.pop(fi)
                        if call_start > fs:
                            free_windows.insert(fi, (fs, call_start))
                        if call_end < fe:
                            free_windows.insert(fi + (1 if call_start > fs else 0), (call_end, fe))
                        placed = True
                        break
            if not placed and free_windows:
                # Fallback: place in first free window
                fs, fe = free_windows[0]
                call_end = min(fs + timedelta(minutes=call_dur), fe)
                if (call_end - fs).total_seconds() / 60 >= MIN_PACKAGE_MIN:
                    rows.append({
                        "Wochentag":    dow_de,
                        "Datum":        date.strftime("%d.%m.%Y"),
                        "Von":          fs.strftime("%H:%M"),
                        "Bis":          call_end.strftime("%H:%M"),
                        "Beschreibung": pkg["description"],
                        "_start":       fs,
                    })
                    if call_end < fe:
                        free_windows[0] = (call_end, fe)
                    else:
                        free_windows.pop(0)

        # ── Recalculate free time after calls ────────────────────────────
        total_free_min = sum(
            (fe - fs).total_seconds() / 60 for fs, fe in free_windows
        )

        if not fill_pkgs or total_free_min <= 0:
            continue

        # ── Allocate durations proportionally ────────────────────────────
        total_weight = sum(p["weight"] for p in fill_pkgs)
        if total_weight <= 0:
            continue

        allocated: list[tuple[dict, float]] = []
        for pkg in fill_pkgs:
            raw_min = pkg["weight"] / total_weight * total_free_min
            alloc = max(MIN_PACKAGE_MIN, round(raw_min / MIN_PACKAGE_MIN) * MIN_PACKAGE_MIN)
            allocated.append((pkg, alloc))

        # Normalize so total allocated == total_free_min
        total_alloc = sum(a for _, a in allocated)
        if total_alloc != total_free_min and total_alloc > 0:
            scale = total_free_min / total_alloc
            allocated = [(pkg, max(MIN_PACKAGE_MIN,
                                   round(a * scale / MIN_PACKAGE_MIN) * MIN_PACKAGE_MIN))
                         for pkg, a in allocated]
            # Final pass: adjust last item to absorb rounding remainder
            total_alloc = sum(a for _, a in allocated)
            diff = total_free_min - total_alloc
            if diff != 0 and allocated:
                pkg_last, a_last = allocated[-1]
                allocated[-1] = (pkg_last, max(MIN_PACKAGE_MIN, a_last + diff))

        # ── Sequentially place into free windows ─────────────────────────
        window_idx = 0
        cursor = free_windows[0][0] if free_windows else day_start

        for pkg, alloc_min in allocated:
            remaining = alloc_min
            while remaining > 0 and window_idx < len(free_windows):
                ws, we = free_windows[window_idx]
                if cursor < ws:
                    cursor = ws
                avail = (we - cursor).total_seconds() / 60
                if avail <= 0:
                    window_idx += 1
                    continue

                use = min(remaining, avail)
                end = cursor + timedelta(minutes=use)

                if use >= MIN_PACKAGE_MIN:
                    rows.append({
                        "Wochentag":    dow_de,
                        "Datum":        date.strftime("%d.%m.%Y"),
                        "Von":          cursor.strftime("%H:%M"),
                        "Bis":          end.strftime("%H:%M"),
                        "Beschreibung": pkg["description"],
                        "_start":       cursor,
                    })

                cursor = end
                remaining -= use
                if cursor >= we:
                    window_idx += 1

    # Sort all rows by date + start time
    rows.sort(key=lambda r: (r["Datum"], r["_start"]))

    # Merge consecutive rows with identical descriptions on the same date
    merged: list[dict] = []
    for r in rows:
        if (merged
            and merged[-1]["Datum"] == r["Datum"]
            and merged[-1]["Beschreibung"] == r["Beschreibung"]
            and merged[-1]["Bis"] == r["Von"]):
            merged[-1]["Bis"] = r["Bis"]
        else:
            merged.append(r)
    rows = merged

    # Strip internal keys, drop zero-duration rows
    clean = []
    for r in rows:
        r.pop("_start", None)
        if r["Von"] < r["Bis"]:
            clean.append(r)

    return clean


def _compute_free_windows(day_start: datetime, day_end: datetime,
                          reserved: list[tuple[datetime, datetime]]) -> list[tuple[datetime, datetime]]:
    """Compute free time windows between reserved blocks within [day_start, day_end]."""
    free = [(day_start, day_end)]
    for rs, re in sorted(reserved, key=lambda x: x[0]):
        new_free = []
        for fs, fe in free:
            if fe <= rs or fs >= re:
                new_free.append((fs, fe))
            else:
                if fs < rs:
                    new_free.append((fs, rs))
                if fe > re:
                    new_free.append((re, fe))
        free = new_free
    return free

# ── Output ────────────────────────────────────────────────────────────────────

def print_table(rows: list[dict]) -> None:
    header = ["Wochentag", "Datum", "Von", "Bis", "Beschreibung"]
    widths = [12, 12, 6, 6, 80]
    sep = "─" * (sum(widths) + len(widths) * 3 + 1)

    print(sep)
    print("  ".join(h.ljust(w) for h, w in zip(header, widths)))
    print(sep)
    for row in rows:
        desc = row["Beschreibung"]
        # wrap description
        first = True
        while desc:
            chunk, desc = desc[:widths[-1]], desc[widths[-1]:]
            if first:
                line = "  ".join([
                    row["Wochentag"].ljust(widths[0]),
                    row["Datum"].ljust(widths[1]),
                    row["Von"].ljust(widths[2]),
                    row["Bis"].ljust(widths[3]),
                    chunk.ljust(widths[4]),
                ])
                first = False
            else:
                line = "  ".join([" " * w for w in widths[:4]] + [chunk])
            print(line)
    print(sep)

def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["Wochentag", "Datum", "Von", "Bis", "Beschreibung"]
    col_widths = [14, 14, 7, 7, 80]

    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 18

    alt_fill = PatternFill("solid", fgColor="D9E1F2")
    for row_idx, row in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col, value=row.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill
        # auto height: 15pt per line in description col
        desc = row.get("Beschreibung", "")
        lines = max(1, desc.count("\n") + 1)
        ws.row_dimensions[row_idx].height = max(15, lines * 15)

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"✅  Written to {path}")

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    args  = [a for a in sys.argv[1:] if not a.startswith("--")]
    to_xlsx = "--csv" in flags or "--xlsx" in flags

    # Optional explicit output path: --out /path/to/file.xlsx
    out_path_arg: str | None = None
    for i, a in enumerate(sys.argv):
        if a == "--out" and i + 1 < len(sys.argv):
            out_path_arg = sys.argv[i + 1]
            to_xlsx = True
            break

    now  = datetime.now()
    year = int(args[0]) if len(args) > 0 else now.isocalendar().year
    week = int(args[1]) if len(args) > 1 else now.isocalendar().week

    print(f"\nLoading week {week}/{year}…")
    events = load_week(year, week)
    print(f"  {len(events)} events loaded")

    manual_entries = extract_manual_entries(events)
    print(f"  {len(manual_entries)} manual entries")

    slots = split_slots(events)
    print(f"  {len(slots)} raw slots")
    aggs  = [a for a in (aggregate(s) for s in slots) if a]
    print(f"  {len(aggs)} meaningful slots (≥{MIN_SLOT_MINUTES}min)")

    dates = sorted({a["date"] for a in aggs} | {e["date"] for e in manual_entries})
    all_commits = all_commits_for_dates(dates)
    print(f"  {len(all_commits)} commits across {len(dates)} day(s)")
    if all_commits:
        attach_commits(aggs, all_commits)

    packages = harvest_packages(aggs, manual_entries)
    total_pkgs = sum(len(v) for v in packages.values())
    print(f"  {total_pkgs} work packages across {len(packages)} day(s)")

    rows = make_rows(aggs, manual_entries, packages)

    print()
    print_table(rows)

    if to_xlsx:
        out = Path(out_path_arg) if out_path_arg else Path(__file__).parent / f"report_KW{week:02d}_{year}.xlsx"
        write_xlsx(rows, out)
